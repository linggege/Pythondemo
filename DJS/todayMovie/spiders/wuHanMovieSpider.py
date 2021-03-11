# -*- coding: utf-8 -*-
import scrapy
from todayMovie.items import TodaymovieItem
import re
import xlrd
import xlwt
import xlutils
from xlutils import copy    #xlutils中导入copy
from openpyxl import load_workbook

class WuhanmoviespiderSpider(scrapy.Spider):
    name = 'wuHanMovieSpider'
    allowed_domains = ['http://basic.10jqka.com.cn']
    # start_urls = ['http://basic.10jqka.com.cn/600519/worth.html']
    start_urls = []
    data = xlrd.open_workbook('股票分析1.xlsm')  # 打开xls文件
    table = data.sheets()[0]  # 打开第一张表
    nrows = table.nrows  # 获取表的行数
    # nrows = 4 #test
    for i in range(nrows):  # 循环逐行打印
        if i == 0:  # 跳过第一行
            continue
        # print(table.row_values(i)[:10])  # 取前十列
        stockCode = table.row_values(i)[:10][1][2:]
        print(stockCode)
        start_urls.append('http://basic.10jqka.com.cn/' + stockCode + '/worth.html')
        start_urls.append('http://basic.10jqka.com.cn/' + stockCode + '/equity.html')
    # start_urls.append('http://basic.10jqka.com.cn/' + "002027" + '/equity.html')

    def parse(self, response):
        self.log('A response from %s just arrived!' % response.url)
        data = xlrd.open_workbook('股票分析1.xlsm')  # 打开xls文件
        table = data.sheets()[0]  # 打开第一张表
        nrows = table.nrows  # 获取表的行数
        for i in range(nrows):  # 循环逐行打印
            if i == 0:  # 跳过第一行
                continue
            # print(table.row_values(i)[:10])  # 取前十列
            stockCode = table.row_values(i)[:10][1][2:]
            # print(stockCode)
            if stockCode in response.url:
                break
        bequity = False
        equity = "equity"
        if equity in response.url:
            bequity = True
        print(bequity)

        if bequity:
            print("bequity true stockCode:" + stockCode)
            subSelector = response.xpath('//tr/th[text()= "A股总股本(股)"]/following-sibling::td[1]/text()').extract()[0]
            strEquity = subSelector[:-1]
            print(strEquity)
            workbook1 = load_workbook('股票分析1.xlsm', keep_vba=True)
            sheet = workbook1['统计']
            sheet.cell(i + 1, 7).value = float(strEquity)
            workbook1.save('股票分析1.xlsm')  # 保存修改
        else:
            print("bequity false stockCode:" + stockCode)
            subSelector = response.xpath('//div[@id="yjycData" and @class="none"]/text()').extract()[0]
            # print(type(subSelector))
            # print(subSelector)
            yclr = subSelector.replace('],[', '|')
            yclr = yclr.replace('\"', '')
            yclr = yclr[2:len(yclr)-2]
            yclr = yclr.split("|")
            # print(yclr)
            # print(type(yclr))
            yclr = yclr[-2]  # 取倒数第二年预测年的数据
            # yclr = yclr[-1]#取最后一个预测年的数据
            # print("list[-1]：" + yclr)
            yclr = yclr.split(",")
            yclr = yclr[2]
            print(yclr)

            #倒数第二年的预测每股净资产
            ycjzc = response.xpath('//div[@class="pr" and @style="z-index:180"]//span/text()').extract()[0]
            # 倒数第一年的预测每股净资产
            # ycjzc = response.xpath('//div[@class="pr" and @style="z-index:179"]//span/text()').extract()[0]
            print(ycjzc)

            workbook1 = load_workbook('股票分析1.xlsm', keep_vba=True)
            sheet = workbook1['统计']
            sheet.cell(i+1, 6).value = float(yclr)
            sheet.cell(i+1, 10).value = float(ycjzc)
            workbook1.save('股票分析1.xlsm')  # 保存修改


            # new_book = copy.copy(data)
            # # 然后用xlutils里面的copy功能，复制一个excel
            # sheet = new_book.get_sheet(0)  # 获取sheet页
            # sheet.write(i, 23, yclr)  # 修改第0行，第一列
            # sheet.write(i, 24, ycjzc)  # 修改第一行，第一列
            # new_book.save('股票分析1.xls')


            item = TodaymovieItem()
            item['yclr'] = yclr
            item['ycjzc'] = ycjzc
            return item
