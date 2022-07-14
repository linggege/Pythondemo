print('地址保存到地址信息表.xlsx')
import addressparser
import pandas as pd
from openpyxl import load_workbook
from openpyxl import workbook
wb = load_workbook('地址信息表.xlsx')
ws = wb.active
#从第一行遍历到最后一行
for row in range(2, ws.max_row + 1):
#取出地址信息那一栏
    cell = ws.cell(row = row, column = 1)
    location_str = [cell.value]
    df = addressparser.transform(location_str)
    #print(df)
    #将省份信息填写到省份那一栏，column的数字是对应列d 数字。
    ws.cell(row = row, column = 2).value = df.iat[0,0]
    #将城市信息填写到城市那一栏
    ws.cell(row = row, column = 3).value = df.iat[0,1]
    #将区域信息填写到区域那一栏
    ws.cell(row = row, column = 4).value = df.iat[0,2]
    ##将详细信息填写到详细那一栏
    ws.cell(row = row, column = 5).value = df.iat[0,3]
#最后保存文件
wb.save('地址信息表.xlsx')
excel_data = pd.read_excel('地址信息表.xlsx')
data = pd.DataFrame(excel_data, columns=['省', '市', '县','详细地址'])
print(data)%
