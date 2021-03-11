# -*- coding: utf-8 -*-
import re
import requests
import json
from addict import Dict
import numpy as np
import datetime  #导入日期时间模块
from dateutil.relativedelta import relativedelta
import xlrd
import xlwt
import xlutils
#from xlutils import copy  #xlutils中导入copy
from openpyxl import load_workbook

now = datetime.date.today()  #当前日期
today = now.strftime('%Y-%m-%d')
last = now - relativedelta(years=+5)  #五年前
lastyear = last.strftime('%Y-%m-%d')

cookies = {
    'PHPSESSID': 'd8574l1fn48a0h953usrfm2fc0',
    'Hm_lvt_210e7fd46c913658d1ca5581797c34e3': '1582793990,1584339671',
    'Hm_lpvt_210e7fd46c913658d1ca5581797c34e3': '1584339813',
}

headers = {
    'Connection': 'keep-alive',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'DNT': '1',
    'X-Requested-With': 'XMLHttpRequest',
    'User-Agent':
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'http://www.dashiyetouzi.com',
    'Referer':
    'http://www.dashiyetouzi.com/tools/compare/historical_valuation.php',
    'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
}


def p_data(stock_type, stockid, fromdate, todate):  #返回最小值，平均值，中位数的元组
    data = {
        'report_type': stock_type,
        'report_stock_id': stockid,
        'from_date': fromdate,
        'to_date': todate
    }
    r = requests.post(
        'http://www.dashiyetouzi.com/tools/compare/historical_valuation_data.php',
        headers=headers,
        cookies=cookies,
        data=data,
        verify=False)

    data1 = Dict(json.loads(r.text))
    list1 = []
    for rows in data1['list']:
        list1.append(rows[1])
    # lens = len(list1)  #长度
    mins = min(list1)  #最小值
    maxs = max(list1)  #最大值
    means = np.mean(list1)  #均值
    medians = np.median(list1)  #中位数
    #print('--------------')
    if stock_type == 'pb':
        print('PB最小值:', mins)
        print('PB平均值:', means)
        print('PB中位数:', medians)
    else:
        print('PETTM最小值:', mins)
        print('PETTM平均值:', means)
        print('PETTM中位数:', medians)

    return mins, means, medians


############################
data = xlrd.open_workbook('股票分析1.xlsm')  # 打开xls文件
table = data.sheets()[0]  # 打开第一张表
nrows = table.nrows  # 获取表的行数
# nrows = 4 #test
for i in range(nrows):  # 循环逐行打印
    if i == 0:  # 跳过第一行
        continue
    # print(table.row_values(i)[:10])  # 取前十列
    stockCode = table.row_values(i)[:10][1][2:]
    stockcode1 = table.row_values(i)[:10][1]
    url = "http://hq.sinajs.cn/list=" + stockcode1
    r = requests.get(url)
    result = re.findall(".*\=\"(.*)\,.*", r.text)
    t = result[0].split(',')
    stockname = t[0]
    print('==========')
    print("{name}:{id}".format(name=stockname, id=stockCode))
    a = p_data('pb', stockCode, lastyear, today)
    b = p_data('pettm', stockCode, lastyear, today)
    workbook1 = load_workbook('股票分析1.xlsm', keep_vba=True)
    sheet = workbook1['统计']
    sheet.cell(i + 1, 4).value = float(a[2])
    sheet.cell(i + 1, 5).value = float(a[0])
    sheet.cell(i + 1, 8).value = float(b[2])
    sheet.cell(i + 1, 9).value = float(a[0])
    workbook1.save('股票分析1.xlsm')  # 保存修改
